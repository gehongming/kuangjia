import json
from openpyxl import load_workbook
class Case:
    '''
    测试用例类，每个测试用例就是它的一个实例
    '''
    def __init__(self):
        self.case_id = None
        self.title = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None
class DoExcel:

    def __init__(self,bookname,sheetname):
        self.bookname=bookname
        self.workbook = load_workbook(bookname)
        self.sheet=self.workbook[sheetname]

    def get_cases(self):
        max_row = self.sheet.max_row
        cases = []  # 列表，用于存放所有用例
        for i in range(2, max_row + 1):
            case = Case()   # 实例化一个case类
            case.case_id = self.sheet.cell(i, 1).value
            case.title = self.sheet.cell(i, 2).value
            case.url = self.sheet.cell(i, 3).value
            case.data = self.sheet.cell(i, 4).value
            case.method = self.sheet.cell(i, 5).value
            case.expected = self.sheet.cell(i, 6).value
            case.actual = self.sheet.cell(i, 7).value
            case.result = self.sheet.cell(i, 8).value
            case.sql = self.sheet.cell(i, 9).value
            if case is not None:
                cases.append(case)
        return cases

    def write_back(self,row,column,value):
        self.sheet.cell(row,column).value = value
        self.workbook.save(self.bookname)
        self.workbook.close()
if __name__ == '__main__':
    import contants

    a = DoExcel(contants.case_file, 'register')
    print(a.get_cases())