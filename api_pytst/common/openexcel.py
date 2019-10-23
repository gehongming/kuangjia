import json
from openpyxl import load_workbook



class DoExcel:
    def __init__(self,excel,sheet):

            self.excel=excel #excel名称
            self.sheet=sheet   #表单名称
    def read(self):
        login_datas = []
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单
        for i in range (2,sheet.max_row+1):  #获取最大行
            user={
                'case_id': sheet.cell(i, 1).value,
                'title': sheet.cell(i, 2).value,
                'url': sheet.cell(i, 3).value,
                'data': sheet.cell(i, 4).value,
                'method': sheet.cell(i, 5).value,
                'expected':sheet.cell(i,6).value,
                'result':sheet.cell(i,8).value,
                'check_sql':sheet.cell(i,9).value
            }
            login_datas.append(user)

        wb.close()
        return login_datas
    def write(self,row,actual,result):
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单
        sheet.cell(row, 8, result)
        sheet.cell(row,7,actual)
        wb.save(self.excel)
        wb.close()



if __name__ == '__main__':

    import contants
    a=DoExcel(contants.case_file,'verifiedUserAuth')
    print(a.read())





















