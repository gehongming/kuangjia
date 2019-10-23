import json
from openpyxl import load_workbook



class DoExcel:
    def __init__(self,excel,sheet):

            self.excel=excel #excel名称
            self.sheet=sheet   #表单名称
    def read(self):
        login_datas = []
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单
        for i in range (2,sheet.max_row+1):  #获取最大行
            user={
                'url': sheet.cell(i, 3).value,
                'method': sheet.cell(i, 5).value,
                'data': json.loads(sheet.cell(i, 4).value),
                'expected':sheet.cell(i,6).value,
                'case_id':sheet.cell(i,1).value,
                'title':sheet.cell(i,2).value,
                'result':sheet.cell(i,8).value,
                'check_sql':sheet.cell(i,9).value
            }
            login_datas.append(user)

        wb.close()
        return login_datas
    def write(self,row,actual,result):
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单
        sheet.cell(row, 8, result)
        sheet.cell(row,7,actual)
        wb.save(self.excel)
        wb.close()

if __name__=='__main__':
    import contants
    do_excel=DoExcel(contants.case_file,'invest')
    cases=do_excel.read()
    print(cases)
    # import api_cookies
    # http_request=api_cookies.HttpCookies()
    # for case in cases:
    #         resp=http_request.http_request(case['url'],case['data'],case['method'])
    #         print(resp.status_code)
    #         print(resp.text)  # 响应文本
    #         resp_dict = resp.json()  # 返回字典
    #         print(resp_dict)
    #         actual=resp.text
    #         if actual==case['expected']:
    #             do_excel.write(case['case_id']+1,actual,'PASS')
    #         else:
    #             do_excel.write(case['case_id']+ 1, actual,'FAIL')
    # #



















